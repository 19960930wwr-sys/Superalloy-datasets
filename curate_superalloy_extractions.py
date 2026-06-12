from __future__ import annotations

import ast
import json
import math
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont


BASE_DIR = Path(__file__).resolve().parents[1]
OUT_DIR = Path(__file__).resolve().parent
TABLE_DIR = OUT_DIR / "tables"
FIG_DIR = OUT_DIR / "figures"
LOG_DIR = OUT_DIR / "logs"

EXTRACTION_DIRS = {
    "ext-1": "tensile_mechanical",
    "ext-2": "physical_mechanical_other",
    "ext-3": "microstructure_creep_fatigue_or_mixed",
    "ext-4": "fracture_elastic_thermal",
}

ELEMENTS = [
    "H", "B", "C", "N", "O", "F", "Na", "Mg", "Al", "Si", "P", "S", "Cl", "Ar",
    "K", "Ca", "Sc", "Ti", "V", "Cr", "Mn", "Fe", "Co", "Ni", "Cu", "Zn", "Ge",
    "As", "Sr", "Y", "Zr", "Nb", "Mo", "Pd", "Ag", "Cd", "Sn", "Sb", "Te", "Ba",
    "La", "Ce", "Nd", "Er", "Hf", "Ta", "W", "Re", "Os", "Ir", "Pt", "Pb", "Bi",
    "Ru",
]

COMMON_FIELDS = [
    "superalloy name",
    "sample name",
    "composition unit",
    "distinguishing factor",
    "synthesis and processing routes",
    "test route/condition",
    "full compositions",
]

EMPTY_MARKERS = {
    "",
    "none",
    "nan",
    "null",
    "n/a",
    "na",
    "not specified",
    "not available",
    "[]",
    "['none']",
    "['None']",
    "[None]",
}

TARGET_SYNONYMS = {
    "gauge length": ["gauge", "gage", "length"],
    "strain rate": ["strain rate", "deformation rate", "dot", "epsilon", "varepsilon"],
    "tensile strength": ["tensile", "uts", "ultimate"],
    "yield strength": ["yield", "ys", "proof stress"],
    "total elongation": ["total elongation", "elongation", "fracture strain", "ductility"],
    "uniform elongation": ["uniform elongation"],
    "gamma prime solvus temperature": ["solvus", "gamma", "phase stability"],
    "density": ["density", "relative density", "rho", "porosity"],
    "liquidus temperature": ["liquidus"],
    "solidus temperature": ["solidus"],
    "hardness": ["hardness", "vickers", "microhardness", "macrohardness", "hrc", "hv"],
    "creep life": ["creep life", "rupture life", "time to rupture", "creep"],
    "oxidation gain": ["oxidation", "mass gain", "weight gain"],
    "compressive strength": ["compressive", "compression", "yield stress", "flow stress"],
    "gamma prime volume fraction": ["volume fraction", "gamma", "precipitate", "recrystallization"],
    "gamma prime size": ["size", "diameter", "particle", "precipitate"],
    "gamma prime morphology": ["morphology", "cuboidal", "spherical", "shape"],
    "phases present": ["phase", "carbide", "tcp", "laves", "gamma", "fcc"],
    "grain size": ["grain size", "crystallite size", "grain"],
    "creep strain rate": ["creep rate", "strain rate", "minimum creep"],
    "stress rupture strength": ["rupture strength", "stress rupture"],
    "fatigue life": ["fatigue", "cycles", "life"],
    "fracture toughness": ["fracture toughness", "toughness", "kic", "k1c"],
    "crack growth rate": ["crack growth", "da/dn", "growth rate"],
    "youngs modulus": ["young", "elastic modulus", "elasticity modulus", "modulus", " e"],
    "shear modulus": ["shear modulus", "modulus", " g"],
    "thermal conductivity": ["thermal conductivity", "conductivity"],
    "thermal expansion coefficient": ["thermal expansion", "expansion coefficient", "cte", "expansivity"],
    "specific heat capacity": ["specific heat", "heat capacity", "cv", "cp"],
}

POSSIBLE_NON_SUPERALLOY_TERMS = [
    "ti-6al-4v",
    "ti6al4v",
    "copper",
    "cu ",
    "steel",
    "aluminum alloy",
    "al alloy",
    "titanium alloy",
]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() in EMPTY_MARKERS else text


def slug(text: str) -> str:
    text = text.replace("γ′", "gamma prime").replace("γ'", "gamma prime")
    text = text.replace("γ", "gamma").replace("′", "'").replace("’", "'")
    text = re.sub(r"[^0-9A-Za-z]+", "_", text).strip("_").lower()
    return re.sub(r"_+", "_", text)


def canonical_property(prop: str) -> str:
    s = slug(prop)
    s = s.replace("young_s", "youngs")
    s = s.replace("gamma_prime_solvus_temperature", "gamma prime solvus temperature")
    s = s.replace("gamma_prime_volume_fraction", "gamma prime volume fraction")
    s = s.replace("gamma_prime_size", "gamma prime size")
    s = s.replace("gamma_prime_morphology_cuboidal_spherical", "gamma prime morphology")
    s = s.replace("creep_strain_rate_minimum_creep_rate", "creep strain rate")
    s = s.replace("_", " ")
    return s.strip()


def normalize_composition_unit(unit: str) -> str:
    u = clean_text(unit).lower().replace(" ", "")
    if not u:
        return ""
    if u in {"wt.%", "wt%", "wtpercent", "weight%", "mass%"}:
        return "wt.%"
    if u in {"at.%", "at%", "atomic%"}:
        return "at.%"
    return clean_text(unit)


def parse_number(value: Any) -> tuple[float | None, int, str]:
    text = clean_text(value)
    if not text:
        return None, 0, ""
    text = text.replace(",", "")
    nums = re.findall(r"[-+]?(?:\d+\.\d+|\d+|\.\d+)(?:[eE][-+]?\d+)?", text)
    if not nums:
        return None, 0, "non_numeric"
    try:
        first = float(nums[0])
    except ValueError:
        return None, len(nums), "non_numeric"
    kind = "single" if len(nums) == 1 else "multiple_or_range"
    return first, len(nums), kind


def standardize_unit_and_value(prop: str, value: Any, unit: Any) -> tuple[float | None, str, str]:
    raw_unit = clean_text(unit)
    unit_l = raw_unit.lower().replace(" ", "")
    num, _, _ = parse_number(value)
    canon = canonical_property(prop)
    if num is None:
        return None, "", ""

    def unit_has(*needles: str) -> bool:
        return any(n in unit_l for n in needles)

    if "temperature" in canon:
        if unit_has("k") and not unit_has("°c", "celsius"):
            return num - 273.15, "degC", "converted_from_K"
        if unit_has("°c", "c", "℃"):
            return num, "degC", ""
    if any(k in canon for k in ["strength", "stress rupture"]):
        if unit_has("gpa"):
            return num * 1000.0, "MPa", "converted_from_GPa"
        if unit_has("mpa", "n/mm"):
            return num, "MPa", ""
    if canon in {"youngs modulus", "shear modulus"}:
        if unit_has("mpa"):
            return num / 1000.0, "GPa", "converted_from_MPa"
        if unit_has("gpa"):
            return num, "GPa", ""
    if "elongation" in canon or "volume fraction" in canon or canon == "density":
        if unit_has("%"):
            return num, "%", ""
    if canon == "density":
        if unit_has("g/cm", "g·cm", "gcm"):
            return num, "g/cm3", ""
        if unit_has("kg/m", "kg·m"):
            return num / 1000.0, "g/cm3", "converted_from_kg_m3"
    if canon in {"creep life", "fatigue life"}:
        if unit_l in {"h", "hr", "hrs", "hour", "hours"} or "hour" in unit_l:
            return num, "h", ""
        if unit_l in {"s", "sec", "seconds"}:
            return num / 3600.0, "h", "converted_from_s"
        if "cycle" in unit_l:
            return num, "cycles", ""
    if canon in {"grain size", "gamma prime size", "gauge length"}:
        if unit_has("nm"):
            return num / 1000.0, "um", "converted_from_nm"
        if unit_has("um", "µm", "μm"):
            return num, "um", ""
        if unit_has("mm"):
            return num * 1000.0 if canon != "gauge length" else num, "um" if canon != "gauge length" else "mm", ""
    if canon == "strain rate":
        if unit_has("s-1", "s^-1", "s⁻¹"):
            return num, "s^-1", ""
    if canon == "thermal conductivity":
        if unit_has("w/(m", "w/m", "w·m"):
            return num, "W/m/K", ""
    if canon == "specific heat capacity":
        if unit_has("j/(g", "j/g"):
            return num * 1000.0, "J/kg/K", "converted_from_J_g_K"
        if unit_has("j/(kg", "j·kg", "j/kg"):
            return num, "J/kg/K", ""
    return num, raw_unit, "raw_unit_retained"


def parse_full_compositions(raw: Any) -> dict[str, Any]:
    text = clean_text(raw)
    if not text:
        return {}
    try:
        parsed = ast.literal_eval(text)
    except (SyntaxError, ValueError):
        return {}
    if isinstance(parsed, dict):
        return parsed
    return {}


def source_doi(path: Path) -> str:
    stem = path.stem
    return stem.replace("-", "/", 1) if stem.startswith("10.") and "-" in stem else stem


def get_by_header(row_map: dict[str, list[Any]], header: str) -> str:
    for value in row_map.get(header, []):
        text = clean_text(value)
        if text:
            return text
    return ""


def build_row_map(header: list[str], row: tuple[Any, ...]) -> dict[str, list[Any]]:
    mapped: dict[str, list[Any]] = defaultdict(list)
    for idx, h in enumerate(header):
        value = row[idx] if idx < len(row) else None
        mapped[h].append(value)
    return mapped


def discover_property_bases(header: list[str]) -> list[str]:
    bases = []
    header_set = set(header)
    for h in header:
        if h in {"superalloy name", "sample name"}:
            continue
        if h.endswith(" name"):
            base = h[:-5]
            if f"{base} value" in header_set or f"{base} unit" in header_set:
                bases.append(base)
    return bases


def property_name_mismatch(prop: str, name: str) -> bool:
    text = clean_text(name).lower()
    if not text:
        return False
    canon = canonical_property(prop)
    needles = TARGET_SYNONYMS.get(canon, [canon])
    return not any(n.lower() in text for n in needles)


def append_quality_flags(
    common: dict[str, Any],
    element_values: dict[str, str],
    property_issues: list[str],
) -> list[str]:
    flags: list[str] = []
    alloy = clean_text(common.get("superalloy_name", ""))
    sample = clean_text(common.get("sample_name", ""))
    comp_unit = clean_text(common.get("composition_unit_normalized", ""))
    nums = []
    for v in element_values.values():
        num, _, _ = parse_number(v)
        if num is not None:
            nums.append(num)
    if not alloy:
        flags.append("missing_alloy_name")
    if not nums:
        flags.append("missing_composition_values")
        if alloy:
            flags.append("alloy_name_only_no_composition")
    if comp_unit and nums:
        total = sum(nums)
        if total < 50:
            flags.append("partial_composition_sum_lt50")
        elif total < 90:
            flags.append("composition_sum_lt90")
        elif total > 110:
            flags.append("composition_sum_gt110")
    if sample and any(term in sample.lower() for term in POSSIBLE_NON_SUPERALLOY_TERMS):
        flags.append("possible_non_superalloy_sample")
    if alloy and any(term in alloy.lower() for term in POSSIBLE_NON_SUPERALLOY_TERMS):
        flags.append("possible_non_superalloy_alloy_name")
    flags.extend(property_issues)
    return sorted(set(flags))


def read_extraction_dir(ext_dir: Path, category: str) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    wide_rows: list[dict[str, Any]] = []
    long_rows: list[dict[str, Any]] = []
    stats: dict[str, Any] = {
        "category": category,
        "folder": ext_dir.name,
        "files_total": 0,
        "files_with_records": 0,
        "records": 0,
        "empty_files": 0,
        "read_errors": [],
        "header_variants": Counter(),
        "property_bases": Counter(),
    }

    files = sorted(ext_dir.glob("*.xlsx"))
    stats["files_total"] = len(files)
    all_property_bases: set[str] = set()

    for file_idx, path in enumerate(files, 1):
        if file_idx % 1000 == 0:
            print(f"[{ext_dir.name}] read {file_idx}/{len(files)} files; records={stats['records']}", flush=True)
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            iterator = ws.iter_rows(values_only=True)
            header_row = next(iterator, None)
            if not header_row:
                stats["empty_files"] += 1
                wb.close()
                continue
            header = ["" if h is None else str(h).strip() for h in header_row]
            stats["header_variants"][tuple(header)] += 1
            prop_bases = discover_property_bases(header)
            for p in prop_bases:
                all_property_bases.add(p)
                stats["property_bases"][p] += 1

            file_records = 0
            for excel_row_num, row in enumerate(iterator, start=2):
                if not any(clean_text(v) for v in row):
                    continue
                row_map = build_row_map(header, row)
                full_comps = parse_full_compositions(get_by_header(row_map, "full compositions"))
                element_values: dict[str, str] = {}
                for elem in ELEMENTS:
                    value = get_by_header(row_map, elem)
                    if not value and elem in full_comps:
                        value = clean_text(full_comps.get(elem))
                    element_values[elem] = value

                common = {
                    "category": category,
                    "source_folder": ext_dir.name,
                    "doi_or_file_id": source_doi(path),
                    "source_file": path.name,
                    "source_row": excel_row_num,
                    "record_id": f"{ext_dir.name}:{path.stem}:{excel_row_num}",
                    "superalloy_name": get_by_header(row_map, "superalloy name"),
                    "sample_name": get_by_header(row_map, "sample name"),
                    "composition_unit_raw": get_by_header(row_map, "composition unit"),
                    "composition_unit_normalized": normalize_composition_unit(get_by_header(row_map, "composition unit")),
                    "distinguishing_factor": get_by_header(row_map, "distinguishing factor"),
                    "synthesis_and_processing_routes": get_by_header(row_map, "synthesis and processing routes"),
                    "test_route_condition": get_by_header(row_map, "test route/condition"),
                    "full_compositions_raw": get_by_header(row_map, "full compositions"),
                }

                nums = [parse_number(v)[0] for v in element_values.values()]
                nums = [v for v in nums if v is not None]
                common["composition_present_element_count"] = len(nums)
                common["composition_sum_numeric"] = round(sum(nums), 6) if nums else ""

                prop_issues: list[str] = []
                wide = dict(common)
                for elem in ELEMENTS:
                    wide[f"element_{elem}"] = element_values[elem]

                for prop in prop_bases:
                    prop_slug = slug(prop)
                    name = get_by_header(row_map, f"{prop} name")
                    value = get_by_header(row_map, f"{prop} value")
                    unit = get_by_header(row_map, f"{prop} unit")
                    source_fig = get_by_header(row_map, f"{prop} sourced figure")
                    wide[f"{prop_slug}_name"] = name
                    wide[f"{prop_slug}_value"] = value
                    wide[f"{prop_slug}_unit"] = unit
                    wide[f"{prop_slug}_sourced_figure"] = source_fig
                    if any([name, value, unit, source_fig]):
                        pflags: list[str] = []
                        if property_name_mismatch(prop, name):
                            pflags.append("property_name_mismatch")
                            prop_issues.append(f"property_name_mismatch:{canonical_property(prop)}")
                        num, n_count, num_kind = parse_number(value)
                        if value and num is None:
                            pflags.append("property_value_non_numeric")
                        if not value and (name or unit):
                            pflags.append("missing_property_value")
                        std_value, std_unit, conversion_note = standardize_unit_and_value(prop, value, unit)
                        if value and unit and conversion_note == "raw_unit_retained":
                            pflags.append("unit_not_standardized")
                        if value and not unit:
                            pflags.append("missing_property_unit")
                        long_rows.append({
                            **common,
                            "property_target_raw": prop,
                            "property_target_canonical": canonical_property(prop),
                            "property_name_raw": name,
                            "property_value_raw": value,
                            "property_unit_raw": unit,
                            "property_value_numeric_first": num if num is not None else "",
                            "property_numeric_token_count": n_count,
                            "property_numeric_parse_kind": num_kind,
                            "property_standard_value": round(std_value, 8) if std_value is not None and math.isfinite(std_value) else "",
                            "property_standard_unit": std_unit,
                            "unit_conversion_note": conversion_note,
                            "property_sourced_figure": source_fig,
                            "property_flags": ";".join(sorted(set(pflags))),
                        })

                flags = append_quality_flags(common, element_values, prop_issues)
                wide["quality_flags"] = ";".join(flags)
                wide["quality_flag_count"] = len(flags)
                wide_rows.append(wide)
                file_records += 1
                stats["records"] += 1
            if file_records:
                stats["files_with_records"] += 1
            else:
                stats["empty_files"] += 1
            wb.close()
        except Exception as exc:  # keep the pipeline moving and log the file.
            stats["read_errors"].append({"file": path.name, "error": repr(exc)})
    stats["property_bases_all"] = sorted(all_property_bases, key=slug)
    return wide_rows, long_rows, stats


def normalize_wide_columns(df: pd.DataFrame, property_bases: list[str]) -> pd.DataFrame:
    first_cols = [
        "category", "source_folder", "doi_or_file_id", "source_file", "source_row", "record_id",
        "superalloy_name", "sample_name", "composition_unit_raw", "composition_unit_normalized",
        "composition_present_element_count", "composition_sum_numeric", "distinguishing_factor",
        "synthesis_and_processing_routes", "test_route_condition", "full_compositions_raw",
    ]
    element_cols = [f"element_{e}" for e in ELEMENTS]
    prop_cols = []
    for prop in sorted(property_bases, key=slug):
        s = slug(prop)
        prop_cols.extend([f"{s}_name", f"{s}_value", f"{s}_unit", f"{s}_sourced_figure"])
    tail_cols = ["quality_flags", "quality_flag_count"]
    for col in first_cols + element_cols + prop_cols + tail_cols:
        if col not in df.columns:
            df[col] = ""
    return df[first_cols + element_cols + prop_cols + tail_cols]


def write_excel(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            safe = sheet_name[:31]
            df.to_excel(writer, sheet_name=safe, index=False)
            ws = writer.book[safe]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions


def draw_bar_chart(path: Path, title: str, labels: list[str], values: list[float], x_label: str = "count") -> None:
    width, height = 1200, max(420, 80 + 44 * len(labels))
    img = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(img)
    font = ImageFont.load_default()
    draw.text((30, 24), title, fill=(20, 20, 20), font=font)
    max_value = max(values) if values else 1
    left, top = 260, 70
    bar_h, gap = 26, 18
    palette = [(54, 117, 136), (204, 119, 34), (68, 147, 85), (145, 97, 153), (180, 90, 90)]
    for i, (label, value) in enumerate(zip(labels, values)):
        y = top + i * (bar_h + gap)
        draw.text((30, y + 6), str(label)[:36], fill=(40, 40, 40), font=font)
        bar_w = int((width - left - 170) * (value / max_value)) if max_value else 0
        draw.rectangle([left, y, left + bar_w, y + bar_h], fill=palette[i % len(palette)])
        draw.text((left + bar_w + 10, y + 6), f"{value:g}", fill=(40, 40, 40), font=font)
    draw.text((left, height - 32), x_label, fill=(80, 80, 80), font=font)
    img.save(path)


def draw_heatmap(path: Path, title: str, matrix: pd.DataFrame) -> None:
    cell_w, cell_h = 155, 34
    left, top = 260, 80
    width = left + cell_w * len(matrix.columns) + 60
    height = top + cell_h * len(matrix.index) + 70
    img = Image.new("RGB", (width, max(360, height)), "white")
    draw = ImageDraw.Draw(img)
    font = ImageFont.load_default()
    draw.text((30, 24), title, fill=(20, 20, 20), font=font)
    max_v = max([float(v) for v in matrix.to_numpy().flatten()] + [1.0])
    for j, col in enumerate(matrix.columns):
        draw.text((left + j * cell_w + 8, top - 26), str(col)[:20], fill=(40, 40, 40), font=font)
    for i, idx in enumerate(matrix.index):
        y = top + i * cell_h
        draw.text((30, y + 9), str(idx)[:38], fill=(40, 40, 40), font=font)
        for j, col in enumerate(matrix.columns):
            x = left + j * cell_w
            v = float(matrix.loc[idx, col])
            intensity = int(245 - 170 * (v / max_v))
            color = (intensity, 235, 245)
            draw.rectangle([x, y, x + cell_w - 2, y + cell_h - 2], fill=color, outline=(220, 220, 220))
            if v:
                draw.text((x + 8, y + 9), f"{int(v)}", fill=(20, 20, 20), font=font)
    img.save(path)


def build_summaries(all_wide: list[pd.DataFrame], all_long: list[pd.DataFrame], stats_list: list[dict[str, Any]]) -> dict[str, pd.DataFrame]:
    wide = pd.concat(all_wide, ignore_index=True) if all_wide else pd.DataFrame()
    long = pd.concat(all_long, ignore_index=True) if all_long else pd.DataFrame()

    category_summary = pd.DataFrame([
        {
            "source_folder": s["folder"],
            "category": s["category"],
            "files_total": s["files_total"],
            "files_with_records": s["files_with_records"],
            "empty_or_no_record_files": s["empty_files"],
            "wide_records": s["records"],
            "read_error_count": len(s["read_errors"]),
            "header_variant_count": len(s["header_variants"]),
            "property_targets_detected": len(s["property_bases_all"]),
        }
        for s in stats_list
    ])

    if not long.empty:
        property_summary = (
            long.groupby(["category", "property_target_canonical"], dropna=False)
            .agg(
                property_rows=("record_id", "count"),
                rows_with_value=("property_value_raw", lambda x: sum(bool(clean_text(v)) for v in x)),
                rows_with_standard_unit=("property_standard_unit", lambda x: sum(bool(clean_text(v)) for v in x)),
                unit_variants=("property_unit_raw", lambda x: len({clean_text(v) for v in x if clean_text(v)})),
                name_variants=("property_name_raw", lambda x: len({clean_text(v) for v in x if clean_text(v)})),
            )
            .reset_index()
            .sort_values(["category", "property_rows"], ascending=[True, False])
        )
        unit_summary = (
            long.assign(property_unit_raw=long["property_unit_raw"].map(clean_text))
            .query("property_unit_raw != ''")
            .groupby(["category", "property_target_canonical", "property_unit_raw"], dropna=False)
            .size()
            .reset_index(name="count")
            .sort_values(["category", "property_target_canonical", "count"], ascending=[True, True, False])
        )
    else:
        property_summary = pd.DataFrame()
        unit_summary = pd.DataFrame()

    if not wide.empty:
        rows = []
        for _, row in wide.iterrows():
            flags = [f for f in str(row.get("quality_flags", "")).split(";") if f]
            for flag in flags:
                base = flag.split(":", 1)[0]
                rows.append({"category": row["category"], "flag": base})
        flag_summary = pd.DataFrame(rows).groupby(["category", "flag"]).size().reset_index(name="count") if rows else pd.DataFrame()
        composition_summary = (
            wide.groupby("category", dropna=False)
            .agg(
                records=("record_id", "count"),
                records_with_any_composition=("composition_present_element_count", lambda x: sum(float(v or 0) > 0 for v in x)),
                median_present_elements=("composition_present_element_count", "median"),
                records_sum_lt50=("quality_flags", lambda x: sum("partial_composition_sum_lt50" in str(v) for v in x)),
                records_missing_composition=("quality_flags", lambda x: sum("missing_composition_values" in str(v) for v in x)),
            )
            .reset_index()
        )
    else:
        flag_summary = pd.DataFrame()
        composition_summary = pd.DataFrame()

    return {
        "category_summary": category_summary,
        "property_summary": property_summary,
        "unit_summary": unit_summary,
        "quality_flag_summary": flag_summary,
        "composition_summary": composition_summary,
    }


def write_readme(summary: dict[str, pd.DataFrame]) -> None:
    readme = OUT_DIR / "README_整理说明.md"
    lines = [
        "# Superalloy extraction curation outputs",
        "",
        f"Generated at: {datetime.now().isoformat(timespec='seconds')}",
        "",
        "## Automatic checks",
        "",
        "- Missing alloy name, missing composition values, alloy-name-only records.",
        "- Low or abnormal composition totals after numeric parsing.",
        "- Duplicate element columns merged into one canonical element column.",
        "- Property target/name mismatch based on conservative keyword rules.",
        "- Missing property values or units, non-numeric property values, and unstandardized units.",
        "- Possible non-superalloy sample/alloy keywords are flagged for review.",
        "",
        "## Output files",
        "",
        "- `tables/*_wide_cleaned.xlsx`: one record per extracted sample row; original values retained, with added QC fields.",
        "- `tables/*_properties_long.xlsx`: one row per property mention; easier for statistics and plotting.",
        "- `tables/summary_statistics.xlsx`: category, property, unit, composition, and quality-flag summaries.",
        "- `figures/*.png`: draft figures for dataset description and quality assessment.",
        "",
        "## Notes for manuscript use",
        "",
        "The pipeline is intentionally conservative: it flags suspicious records instead of deleting or inferring missing data. "
        "Grade-to-composition completion should be done with a curated alloy designation database if it is needed.",
    ]
    if "category_summary" in summary and not summary["category_summary"].empty:
        lines.extend(["", "## Category summary", ""])
        cols = list(summary["category_summary"].columns)
        lines.append(" | ".join(cols))
        lines.append(" | ".join(["---"] * len(cols)))
        for _, row in summary["category_summary"].iterrows():
            lines.append(" | ".join(str(row[col]) for col in cols))
    readme.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    TABLE_DIR.mkdir(parents=True, exist_ok=True)
    FIG_DIR.mkdir(parents=True, exist_ok=True)
    LOG_DIR.mkdir(parents=True, exist_ok=True)

    all_wide_dfs: list[pd.DataFrame] = []
    all_long_dfs: list[pd.DataFrame] = []
    stats_list: list[dict[str, Any]] = []

    for folder, category in EXTRACTION_DIRS.items():
        ext_dir = BASE_DIR / folder
        print(f"Processing {folder} -> {category}", flush=True)
        wide_rows, long_rows, stats = read_extraction_dir(ext_dir, category)
        stats_list.append(stats)
        property_bases = stats["property_bases_all"]

        wide_df = normalize_wide_columns(pd.DataFrame(wide_rows), property_bases) if wide_rows else normalize_wide_columns(pd.DataFrame(), property_bases)
        long_df = pd.DataFrame(long_rows)
        long_columns = [
            "category", "source_folder", "doi_or_file_id", "source_file", "source_row", "record_id",
            "superalloy_name", "sample_name", "composition_unit_raw", "composition_unit_normalized",
            "composition_present_element_count", "composition_sum_numeric", "distinguishing_factor",
            "synthesis_and_processing_routes", "test_route_condition", "full_compositions_raw",
            "property_target_raw", "property_target_canonical", "property_name_raw", "property_value_raw",
            "property_unit_raw", "property_value_numeric_first", "property_numeric_token_count",
            "property_numeric_parse_kind", "property_standard_value", "property_standard_unit",
            "unit_conversion_note", "property_sourced_figure", "property_flags",
        ]
        for col in long_columns:
            if col not in long_df.columns:
                long_df[col] = ""
        long_df = long_df[long_columns]

        wide_path = TABLE_DIR / f"{folder}_{category}_wide_cleaned.xlsx"
        long_path = TABLE_DIR / f"{folder}_{category}_properties_long.xlsx"
        write_excel(wide_path, {"cleaned_wide": wide_df})
        write_excel(long_path, {"properties_long": long_df})
        wide_df.to_csv(TABLE_DIR / f"{folder}_{category}_wide_cleaned.csv", index=False, encoding="utf-8-sig")
        long_df.to_csv(TABLE_DIR / f"{folder}_{category}_properties_long.csv", index=False, encoding="utf-8-sig")

        all_wide_dfs.append(wide_df)
        all_long_dfs.append(long_df)

    log_payload = []
    for stats in stats_list:
        compact = dict(stats)
        compact["header_variants"] = {str(k): v for k, v in stats["header_variants"].items()}
        compact["property_bases"] = dict(stats["property_bases"])
        log_payload.append(compact)
    (LOG_DIR / "processing_log.json").write_text(json.dumps(log_payload, ensure_ascii=False, indent=2), encoding="utf-8")

    summaries = build_summaries(all_wide_dfs, all_long_dfs, stats_list)
    write_excel(TABLE_DIR / "summary_statistics.xlsx", summaries)
    for name, df in summaries.items():
        df.to_csv(TABLE_DIR / f"{name}.csv", index=False, encoding="utf-8-sig")

    cat = summaries["category_summary"]
    if not cat.empty:
        draw_bar_chart(
            FIG_DIR / "records_by_category.png",
            "Extracted records by performance category",
            cat["category"].tolist(),
            cat["wide_records"].astype(float).tolist(),
            "wide records",
        )
        draw_bar_chart(
            FIG_DIR / "files_with_records_by_category.png",
            "Literature files with at least one extracted record",
            cat["category"].tolist(),
            cat["files_with_records"].astype(float).tolist(),
            "files",
        )

    comp = summaries["composition_summary"]
    if not comp.empty:
        draw_bar_chart(
            FIG_DIR / "missing_composition_by_category.png",
            "Records flagged as missing composition",
            comp["category"].tolist(),
            comp["records_missing_composition"].astype(float).tolist(),
            "records",
        )

    prop = summaries["property_summary"]
    if not prop.empty:
        pivot = prop.pivot_table(
            index="property_target_canonical",
            columns="category",
            values="property_rows",
            aggfunc="sum",
            fill_value=0,
        )
        pivot = pivot.loc[pivot.sum(axis=1).sort_values(ascending=False).index[:30]]
        draw_heatmap(FIG_DIR / "property_record_heatmap_top30.png", "Property records by category", pivot)
        top_units = prop.sort_values("unit_variants", ascending=False).head(20)
        draw_bar_chart(
            FIG_DIR / "unit_variant_count_top20_properties.png",
            "Top properties by unit-name diversity",
            (top_units["category"] + " | " + top_units["property_target_canonical"]).tolist(),
            top_units["unit_variants"].astype(float).tolist(),
            "distinct unit strings",
        )

    flags = summaries["quality_flag_summary"]
    if not flags.empty:
        top_flags = flags.groupby("flag")["count"].sum().sort_values(ascending=False).head(20)
        draw_bar_chart(
            FIG_DIR / "quality_flags_top20.png",
            "Most frequent automatic quality flags",
            top_flags.index.tolist(),
            top_flags.astype(float).tolist(),
            "flagged records",
        )

    write_readme(summaries)
    print(f"Done. Outputs written to: {OUT_DIR}", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
